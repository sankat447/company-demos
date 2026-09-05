#!/usr/bin/env python3
"""
Generate the Bir Festival 2026 data-collection workbook.

One sheet per data entity the app ingests, with typed columns, required-field
markers (*), English/Hindi pairs, dropdown validation for every enum, a sample
row, and header-cell comments describing each field. Human-friendly inputs
(dates as YYYY-MM-DD, times as HH:MM 24h, phones as +91XXXXXXXXXX); the import
step converts dates/times to epoch seconds and looks up cross-sheet ids.

Run:  python3 generate_template.py
Out:  Bir_Festival_2026_Data_Collection.xlsx
"""
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

INK = "17232B"
PINE = "2E5E4E"
SLATE = "3E6B8C"
MARIGOLD = "E8A13D"
PAPER = "F7F8F5"
EXAMPLE = "FCF3E3"
GREY = "5B6B75"

HEADER_FILL = PatternFill("solid", fgColor=PINE)
EXAMPLE_FILL = PatternFill("solid", fgColor=EXAMPLE)
TITLE_FILL = PatternFill("solid", fgColor=INK)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
EX_FONT = Font(italic=True, color=INK)
THIN = Side(style="thin", color="D5DAD3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

YESNO = '"yes,no"'

# A column = (key/label, required, type_hint, example, enum_or_None, help_text)
def col(label, req, typ, example, enum=None, help="" ):
    return dict(label=label, req=req, typ=typ, example=example, enum=enum, help=help)

SHEETS = [
    dict(name="1. Ticket Tiers", color=PINE,
         desc="Festival entry passes shown on the Buy screen.",
         cols=[
            col("id", True, "text", "day-pass", None, "Unique, lowercase-kebab. e.g. day-pass, 3day-pass."),
            col("title_en", True, "text", "Day pass"),
            col("title_hi", True, "text", "दिन का पास"),
            col("price_inr", True, "whole number", 499, None, "Price in INR, no symbols."),
            col("description_en", False, "text", "Valid for one festival day"),
            col("description_hi", False, "text", "एक दिन के लिए मान्य"),
         ]),
    dict(name="2. Venues", color=PINE,
         desc="Places used across schedule, highlights and maps. Fill this first — other sheets reference venue_id.",
         cols=[
            col("id", True, "text", "chogan", None, "Unique, lowercase-kebab. Referenced by venue_id elsewhere."),
            col("name_en", True, "text", "Chogan Ground"),
            col("name_hi", False, "text", "चौगान मैदान"),
            col("latitude", True, "decimal", 32.0322, None, "Decimal degrees."),
            col("longitude", True, "decimal", 76.7185),
            col("notes", False, "text", "Main stage & award ceremony"),
         ]),
    dict(name="3. Cultural-Night Schedule", color=PINE,
         desc="Evening programme (21-23 Nov). Votable rows feed the audience-award voting.",
         cols=[
            col("id", True, "text", "d1-folk", None, "Unique per event."),
            col("date", True, "YYYY-MM-DD", "2026-11-21"),
            col("venue_id", True, "ref Venues.id", "chogan"),
            col("title_en", True, "text", "Folk music of Kangra"),
            col("title_hi", True, "text", "कांगड़ा का लोक संगीत"),
            col("start_time", True, "HH:MM 24h", "18:00"),
            col("end_time", False, "HH:MM 24h", "19:00"),
            col("votable", False, "yes/no", "yes", YESNO, "yes = eligible for the audience-favourite vote."),
            col("category", False, "text", "music", None, "Free text, e.g. music, dance, comedy."),
         ]),
    dict(name="4. Highlight Categories", color=MARIGOLD,
         desc="The top-level tiles in the Highlights hub. Usually 6 rows.",
         cols=[
            col("id", True, "text", "competitions", None, "lowercase-kebab. Referenced by category_id in Highlight Items."),
            col("title_en", True, "text", "Competitions"),
            col("title_hi", True, "text", "प्रतियोगिताएँ"),
            col("icon_emoji", True, "emoji", "🏆", None, "Single emoji shown on the tile."),
            col("sort_order", True, "whole number", 1, None, "Display order 1..N."),
            col("kind", True, "enum", "competition", '"competition,agenda,session,adventure,tour"',
                "competition=judged; agenda=timed line-up; session=class; adventure/tour=slotted booking."),
         ]),
    dict(name="5. Highlight Items", color=MARIGOLD,
         desc="Every bookable/viewable activity: competitions, yoga, pottery, adventure, sightseeing, etc.",
         cols=[
            col("id", True, "text", "chef-local", None, "Unique, lowercase-kebab."),
            col("category_id", True, "ref Categories.id", "competitions"),
            col("title_en", True, "text", "Chef of the Year — Local Food"),
            col("title_hi", True, "text", "शेफ ऑफ द ईयर — स्थानीय भोजन"),
            col("summary_en", True, "text", "Cook the valley's signature dishes"),
            col("summary_hi", True, "text", "घाटी के प्रसिद्ध व्यंजन बनाएं"),
            col("venue_id", True, "ref Venues.id", "chogan"),
            col("dates", True, "YYYY-MM-DD; comma", "2026-11-21,2026-11-22", None, "One or more dates, comma-separated."),
            col("fee_inr", False, "whole number", 200, None, "Blank = free."),
            col("capacity", False, "whole number", 30, None, "Blank = unlimited."),
            col("reg_mode", True, "enum", "register-participation",
                '"register,register-participation,view-only"',
                "register=book a slot; register-participation=competition entry; view-only=info only."),
            col("gate_checked", False, "yes/no", "yes", YESNO, "yes = a scannable badge/pass is issued at the gate."),
            col("guardian_required", False, "yes/no", "no", YESNO, "yes = minors need guardian consent."),
            col("weather_sensitive", False, "yes/no", "no", YESNO, "yes = paused when fly-status is Hold/Closed."),
            col("has_slots", False, "yes/no", "no", YESNO, "yes = also add rows in the 'Highlight Slots' sheet."),
            col("is_competition", False, "yes/no", "yes", YESNO,
                "yes = intake collects gender/needs-lodging/couple fields (feeds lodging)."),
            col("rules_en", False, "text", "Two rounds; local ingredients only"),
            col("rules_hi", False, "text", "दो राउंड; केवल स्थानीय सामग्री"),
            col("eligibility_en", False, "text", "Open to Kangra residents 18+"),
            col("eligibility_hi", False, "text", "कांगड़ा निवासी 18+"),
         ]),
    dict(name="6. Highlight Slots", color=MARIGOLD,
         desc="Time slots for items where has_slots = yes (paragliding, tours, classes).",
         cols=[
            col("item_id", True, "ref Items.id", "paragliding", None, "Must match a Highlight Items id with has_slots=yes."),
            col("slot_id", True, "text", "pg-21-am", None, "Unique per item."),
            col("date", True, "YYYY-MM-DD", "2026-11-21"),
            col("start_time", True, "HH:MM 24h", "09:00"),
            col("label_en", True, "text", "Nov 21 · 09:00 · Pilot group A"),
            col("label_hi", True, "text", "21 नव॰ · 09:00 · पायलट समूह A"),
            col("capacity", False, "whole number", 12),
         ]),
    dict(name="7. Competition Participants", color=PINE,
         desc="Confirmed competition entrants. Feeds registrations AND the lodging pool (gender is lodging-only, never shown on badges).",
         cols=[
            col("reg_id", True, "text", "reg-anita-01", None, "Unique per participant."),
            col("name", True, "text", "Anita Thakur"),
            col("phone", False, "+91XXXXXXXXXX", "+919876500001"),
            col("competition_id", True, "ref Items.id", "him-queen-2026", None, "A Highlight Items id where is_competition=yes."),
            col("gender", True, "enum", "female", '"female,male,other,undisclosed"',
                "Used ONLY for lodging allocation. Never printed on badges or hotel rosters."),
            col("age", False, "whole number", 24),
            col("needs_lodging", True, "yes/no", "yes", YESNO),
            col("lodging_nights", False, "YYYY-MM-DD; comma", "2026-11-21,2026-11-22,2026-11-23",
                None, "Nights they need a room. Required if needs_lodging=yes."),
            col("couple_group_id", False, "text", "", None, "Same value on both partners' rows to keep them together."),
            col("partner_ref", False, "text", "", None, "Partner's reg_id or phone (couples only)."),
            col("status", False, "enum", "confirmed", '"pending,confirmed,waitlist,cancelled"'),
            col("notes", False, "text", "Vegetarian kitchen access"),
         ]),
    dict(name="8. Rooms Inventory", color=SLATE,
         desc="Lodging rooms available to allocate to participants.",
         cols=[
            col("id", True, "text", "r-surya-101", None, "Unique per room."),
            col("hotel_name", True, "text", "Hotel Surya Classic"),
            col("room_label", True, "text", "101"),
            col("type", True, "enum", "twin", '"twin,double,triple,dorm"'),
            col("capacity", True, "whole number", 2),
            col("double_occupancy", False, "yes/no", "no", YESNO, "yes = couple-eligible (double rooms only)."),
            col("available_nights", True, "YYYY-MM-DD; comma", "2026-11-21,2026-11-22,2026-11-23"),
            col("contact_phone", False, "+91XXXXXXXXXX", "+919876511111"),
            col("property_id", False, "text", "", None, "Link to a partner hotel (optional)."),
            col("amenities_note", False, "text", "Attached bath, heater"),
            col("status", True, "enum", "active", '"active,retired"'),
         ]),
    dict(name="9. Volunteers", color=SLATE,
         desc="Volunteer roster. Phone is their app login (Cognito username).",
         cols=[
            col("name", True, "text", "Karan Verma"),
            col("phone", True, "+91XXXXXXXXXX", "+919876522222", None, "Login identity — must be reachable for OTP."),
            col("team", True, "text", "Gate & Access — Team C"),
            col("id_verified", False, "yes/no", "no", YESNO),
            col("notes", False, "text", ""),
         ]),
    dict(name="10. Volunteer Shifts", color=SLATE,
         desc="Shifts per volunteer (link by phone).",
         cols=[
            col("volunteer_phone", True, "ref Volunteers.phone", "+919876522222"),
            col("shift_id", True, "text", "s-21-am"),
            col("date", True, "YYYY-MM-DD", "2026-11-21"),
            col("zone", True, "text", "Chogan Gate 2"),
            col("role", True, "text", "Gate scanning"),
            col("start_time", True, "HH:MM 24h", "08:00"),
            col("end_time", True, "HH:MM 24h", "12:00"),
         ]),
    dict(name="11. Partners - Stalls", color=SLATE,
         desc="Food-street vendors.",
         cols=[
            col("stall_name", True, "text", "Kangra Kitchen"),
            col("vendor_phone", False, "+91XXXXXXXXXX", "+919876533333"),
            col("category", True, "text", "Local food · siddu & dham"),
            col("stage", True, "enum", "approved", '"applied,approved,payment-pending,active,closed"'),
            col("allocation_label", False, "text", "Food Street · Stall F-12"),
            col("fee_inr", False, "whole number", 3500),
            col("paid", False, "yes/no", "no", YESNO),
            col("rules_en", False, "text; use ; between rules", "No single-use plastic; Deposit-return cups"),
            col("rules_hi", False, "text; use ; between rules", "एकल-उपयोग प्लास्टिक नहीं; डिपॉज़िट-वापसी कप"),
         ]),
    dict(name="12. Partners - Hospitality", color=SLATE,
         desc="Partner hotels/homestays (their complimentary-room programme).",
         cols=[
            col("hotel_name", True, "text", "Deodar Homestay"),
            col("manager_phone", False, "+91XXXXXXXXXX", "+919876544444"),
            col("tier", True, "text", "11+ rooms → two complimentary twin rooms, two nights"),
            col("complimentary_rooms", False, "whole number", 2),
            col("total_rooms", False, "whole number", 14),
            col("notes", False, "text", ""),
         ]),
    dict(name="13. Users & Roles", color=INK,
         desc="People who need a privileged login. Role decides which screens they see.",
         cols=[
            col("name", True, "text", "Priya Sharma"),
            col("phone", True, "+91XXXXXXXXXX", "+919876555555", None, "App login (OTP)."),
            col("role", True, "enum", "admin-hospitality",
                '"visitor,partner,volunteer,organiser-lite,admin-hospitality,safety-officer"',
                "visitor=default; partner=stall/hotel; volunteer=scanner/roster; organiser-lite=ops; admin-hospitality=lodging; safety-officer=declares fly-status."),
            col("org_team", False, "text", "Hospitality desk"),
            col("notes", False, "text", ""),
         ]),
    dict(name="14. Fly Status (initial)", color=INK,
         desc="The starting paragliding status. One row. Safety officers change it live in the app.",
         cols=[
            col("state", True, "enum", "flying", '"flying,hold,closed"'),
            col("reason_en", False, "text", "Clear skies over Billing"),
            col("reason_hi", False, "text", "बिलिंग के ऊपर साफ़ आसमान"),
            col("refunds_auto_queued", False, "yes/no", "no", YESNO, "yes = a Hold/Closed auto-queues affected refunds."),
         ]),
    dict(name="15. Configuration", color=INK,
         desc="Deployment/ops values the app needs. Fill the 'value' column. Some are pre-filled defaults.",
         cols=[
            col("config_key", True, "text", "ops.emergencyPhone"),
            col("value", False, "text", "112"),
            col("required", False, "yes/no", "no"),
            col("purpose", False, "text", "SOS fallback number"),
         ],
         prefill=[
            ("ops.emergencyPhone", "112", "no", "SOS fallback dialer number (default 112)"),
            ("passes.issuerKid", "bir-2026-01", "yes", "JWT key id for pass verification (already set by deploy)"),
            ("payments.provider", "", "yes", "razorpay | cashfree"),
            ("payments.providerKeyId", "", "yes", "Public key id for the payment SDK"),
            ("highlights.catalogPath", "", "yes", "CDN path serving the Highlights catalog JSON"),
            ("geo.shuttleEtaPath", "", "no", "Shuttle ETA REST endpoint (optional)"),
            ("ops.lostFoundPath", "", "no", "Lost & found REST endpoint (optional)"),
            ("refunds.autoQueueOnHold", "yes", "yes", "Auto-queue refunds when fly-status = hold/closed"),
         ]),
]


def add_instructions(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 110
    ws["B2"] = "Bir Festival 2026 — Data Collection Workbook"
    ws["B2"].font = Font(bold=True, size=18, color=INK)
    ws["B3"] = "Fill each tab with real festival data, then hand it back for import into the app."
    ws["B3"].font = Font(size=11, color=GREY)
    lines = [
        ("How to use", True),
        ("• One row = one record. Do not rename or reorder the column headers.", False),
        ("• Columns marked with * are required. Grey helper text sits in a comment on each header cell (hover the red corner).", False),
        ("• Row 2 of every sheet is a shaded EXAMPLE — overwrite it or delete it before returning the file.", False),
        ("• Cells with a dropdown arrow only accept the listed values.", False),
        ("", False),
        ("Formats (the import step converts these automatically)", True),
        ("• Dates: YYYY-MM-DD  (e.g. 2026-11-21). Festival days are 21, 22, 23 November 2026.", False),
        ("• Times: HH:MM in 24-hour local time (e.g. 18:00). Times become epoch timestamps on import.", False),
        ("• Phone numbers: +91 then 10 digits (e.g. +919876500001). This is the app login (OTP).", False),
        ("• Lists in one cell: separate with commas (dates) or semicolons (rules).", False),
        ("• IDs: short lowercase-with-hyphens, unique within the sheet. Other sheets reference them (venue_id, category_id, item_id, competition_id, volunteer_phone).", False),
        ("", False),
        ("Fill order (because sheets reference each other)", True),
        ("1) Venues → 2) Ticket Tiers → 3) Schedule → 4) Highlight Categories → 5) Highlight Items → 6) Slots", False),
        ("7) Competition Participants → 8) Rooms → 9) Volunteers → 10) Shifts → 11-12) Partners → 13) Users & Roles → 14) Fly Status → 15) Configuration", False),
        ("", False),
        ("Privacy (please respect)", True),
        ("• Gender is collected ONLY on 'Competition Participants' and is used ONLY for lodging allocation. It must never appear on badges or hotel rosters.", False),
        ("• Collect only what a tab asks for. No ID-card numbers, no medical details in free-text notes.", False),
        ("", False),
        ("Bilingual", True),
        ("• Every _en field needs its _hi (Hindi) counterpart — the app is bilingual and shows Hindi to many users.", False),
    ]
    r = 5
    for text, is_head in lines:
        c = ws.cell(row=r, column=2, value=text)
        if is_head:
            c.font = Font(bold=True, size=12, color=PINE)
        else:
            c.font = Font(size=10.5, color=INK)
            c.alignment = Alignment(wrap_text=True)
        r += 1


def add_sheet(wb, spec):
    ws = wb.create_sheet(spec["name"][:31])
    cols = spec["cols"]
    # Title band (row 1) spanning the columns
    ncol = len(cols)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    t = ws.cell(row=1, column=1, value=f'  {spec["name"].split(". ",1)[-1]} — {spec["desc"]}')
    t.fill = PatternFill("solid", fgColor=spec["color"])
    t.font = Font(bold=True, color="FFFFFF", size=11)
    t.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26

    # Header row (row 2)
    for i, cdef in enumerate(cols, start=1):
        label = cdef["label"] + (" *" if cdef["req"] else "")
        cell = ws.cell(row=2, column=i, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center")
        meta = f'{cdef["typ"]}{" · required" if cdef["req"] else " · optional"}'
        if cdef["help"]:
            meta += f'\n{cdef["help"]}'
        cell.comment = Comment(meta, "template")
        w = max(len(label) + 3, len(str(cdef["example"])) + 3, 12)
        ws.column_dimensions[get_column_letter(i)].width = min(w, 40)

    prefill = spec.get("prefill", [])
    if prefill:
        # Real starter rows instead of a generic example (e.g. Configuration).
        for j, row in enumerate(prefill, start=0):
            for i, val in enumerate(row, start=1):
                c = ws.cell(row=3 + j, column=i, value=val)
                c.border = BORDER
    else:
        # Example row (row 3), shaded, to be overwritten or deleted.
        for i, cdef in enumerate(cols, start=1):
            cell = ws.cell(row=3, column=i, value=cdef["example"])
            cell.fill = EXAMPLE_FILL
            cell.font = EX_FONT
            cell.border = BORDER
        ws.cell(row=3, column=1).comment = Comment(
            "EXAMPLE ROW — overwrite or delete before import.", "template")

    # Enum dropdowns applied down the column
    last = 500
    for i, cdef in enumerate(cols, start=1):
        if cdef["enum"]:
            dv = DataValidation(type="list", formula1=cdef["enum"], allow_blank=True)
            dv.error = "Pick a value from the list."
            dv.errorTitle = "Invalid value"
            ws.add_data_validation(dv)
            colL = get_column_letter(i)
            dv.add(f"{colL}3:{colL}{last}")

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncol)}2"


def main():
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Instructions"
    add_instructions(ws0)
    for spec in SHEETS:
        add_sheet(wb, spec)
    out = "Bir_Festival_2026_Data_Collection.xlsx"
    wb.save(out)
    print(f"wrote {out} with {len(wb.sheetnames)} sheets:")
    for n in wb.sheetnames:
        print("  -", n)


if __name__ == "__main__":
    main()
