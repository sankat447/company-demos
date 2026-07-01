#!/usr/bin/env python
"""Export ALL demo data to docs/seed-data.xlsx — one sheet per table + the derived
analytics the UI shows, for use as a raw-data companion during the demo.

Data is generated deterministically (ensure_seeded + augment_seed, seed=42) so the
workbook matches exactly what the app serves. Reproducible:

    backend/.venv/bin/python docs/build_seed_workbook.py

FOR DEMONSTRATION ONLY — SYNTHETIC DATA. No PHI.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from nychhc_copilot.config import Settings
from nychhc_copilot.scheduling import augment_seed, ensure_seeded, reporting as R
from nychhc_copilot.tools.providers import build_providers

OUT = Path(__file__).parent / "seed-data.xlsx"
FONT = "Calibri"
CLAY = "B05730"
TINT = "F3E2D9"
INK3 = "9B9791"

HFONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
HFILL = PatternFill("solid", fgColor=CLAY)
NOTE = Font(name=FONT, italic=True, size=9, color=INK3)
BASE = Font(name=FONT, size=10)


def _autosize(ws, ncols, rows, header_row):
    for c in range(1, ncols + 1):
        width = len(str(ws.cell(header_row, c).value or ""))
        for r in range(header_row + 1, min(ws.max_row, header_row + 200) + 1):
            v = ws.cell(r, c).value
            if v is not None:
                width = max(width, len(str(v)))
        ws.column_dimensions[get_column_letter(c)].width = min(max(width + 2, 9), 60)


def add_sheet(wb, title, columns, rows, note="", pct_cols=(), num_cols=()):
    ws = wb.create_sheet(title=title[:31])
    if note:
        ws.append([note])
        ws.cell(1, 1).font = NOTE
    ws.append(list(columns))
    hr = ws.max_row
    for c in range(1, len(columns) + 1):
        cell = ws.cell(hr, c)
        cell.font = HFONT
        cell.fill = HFILL
        cell.alignment = Alignment(vertical="center", horizontal="left")
    for row in rows:
        ws.append(["" if v is None else v for v in row])
    for r in range(hr + 1, ws.max_row + 1):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(r, c)
            cell.font = BASE
            if c in pct_cols:
                cell.number_format = "0.0%"
            elif c in num_cols:
                cell.number_format = "#,##0"
    ws.freeze_panes = ws.cell(hr + 1, 1)
    _autosize(ws, len(columns), rows, hr)
    return ws


# ── raw tables (exactly what the backend seeds) ──────────────────────────────
_RAW = [
    ("sched_providers", "Providers", "OBGYN provider roster — the 12 named providers (id, name, credential, specialty, type, hours)."),
    ("sched_patients", "Patients", "Synthetic patients: 15 scripted demo + fill + PT2xxx enrichment. Phones 555-01xx, MRNs SYN-xxxx. No PHI."),
    ("sched_appointments", "Appointments", "Upcoming ~4-week schedule: scripted demo appts (a#) + filler (ax#). Scored live by the KServe no-show model."),
    ("appt_history", "Appt History", "~18-month history corpus (~2,400 rows) — trains the no-show model and drives ASK1/3/4 analytics. outcome = attended / advance_cancel / no_show."),
    ("walkin_daily", "Walk-in Daily", "Per-day walk-in volume (AM/PM) over the last 12 weeks — ASK1 walk-in template signal (Fri is AM-heavy)."),
    ("cycle_log", "Cycle Log", "Per-referral hand-off timings (clerical / scheduling / provider), recent cohort vs prior quarter — ASK3 cycle time."),
    ("sched_pto", "PTO Requests", "Provider leave blocks incl. the scripted Brooks/Wu High-Risk overlap conflict (UC4)."),
    ("pto_queue", "PTO Queue", "Dashboard PTO approval queue with the coverage-gap flag (UC4/ASK2)."),
    ("risk_today", "Risk Panel (UC1)", "At-risk appointments today — Daniel Brooks #1 @ 87%. ~45 rows, red/amber/green with factors."),
    ("roster", "Roster (dashboard)", "Staff roster with PTO balances shown on the dashboard."),
    ("audit_log", "Audit Log (UC6)", "Human-in-the-loop decisions — actor (user + role), decision, outcome, timestamp."),
]


def _readme(wb, index):
    ws = wb.create_sheet("README", 0)
    ws.sheet_view.showGridLines = False
    ws.append(["NYC Health + Hospitals — OBGYN AI Scheduling Assistant"])
    ws.cell(1, 1).font = Font(name=FONT, bold=True, size=16, color=CLAY)
    ws.append(["Demo seed data — all tables + the analytics the UI shows"])
    ws.cell(2, 1).font = Font(name=FONT, size=11, color=INK3)
    ws.append(["⚠  FOR DEMONSTRATION ONLY — SYNTHETIC DATA. No PHI. Names/phones (555-01xx)/MRNs (SYN-xxxx) are fictional."])
    ws.cell(3, 1).font = Font(name=FONT, bold=True, size=10, color="B24A38")
    ws.append([])
    ws.append(["Sheet", "Rows", "What it is"])
    hr = ws.max_row
    for c in range(1, 4):
        ws.cell(hr, c).font = HFONT
        ws.cell(hr, c).fill = HFILL
    for name, count, desc in index:
        ws.append([name, count, desc])
        for c in range(1, 4):
            ws.cell(ws.max_row, c).font = BASE
    ws.append([])
    ws.append(["Raw tables are seeded by the backend (ensure_seeded + augment_seed, seed=42) — identical to the live database."])
    ws.append(["Analytics sheets (KPI Scorecard, Capacity, Cancellations, Cycle Time, …) are computed snapshots of that data — the same numbers the UI renders."])
    for r in range(ws.max_row - 1, ws.max_row + 1):
        ws.cell(r, 1).font = NOTE
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 110
    ws.freeze_panes = "A6"


def build():
    p = build_providers(Settings())
    ensure_seeded(p.aurora)
    augment_seed(p.aurora)
    A = p.aurora

    wb = Workbook()
    wb.remove(wb.active)
    index = []

    # raw tables
    for table, title, note in _RAW:
        try:
            res = A.query(f"SELECT * FROM {table}")
            add_sheet(wb, title, res.columns, res.rows, note=note)
            index.append((title, len(res.rows), note))
        except Exception as e:  # noqa: BLE001
            print(f"skip {table}: {e}")

    # derived analytics (computed snapshots the UI shows)
    b = R.birdseye(A, p.models)

    kpi_rows = [[k["kpi"], k["unit"], k["target"], k["m1"], k["m2"], k["m3"],
                 k["quarter"], k["trend"], k["status"]] for k in b["kpis"]]
    add_sheet(wb, "KPI Scorecard", ["KPI", "Unit", "Target", "Jul", "Aug", "Sep", "Quarter", "Trend", "Status"],
              kpi_rows, note="3-Month Bird's-Eye KPIs vs targets (RAG). Leadership → Reporting tab.")
    index.append(("KPI Scorecard", len(kpi_rows), "3-month KPI scorecard with targets + RAG status (Reporting tab)."))

    cap = b["capacity"]
    cap_rows = [[r["weekday"], r["providers"], r["demand_min"], r["supply_min"], r["utilization"], r["flag"]]
                for r in cap["rows"]]
    cap_rows.append([cap["weekly"]["weekday"], cap["weekly"]["providers"], cap["weekly"]["demand_min"],
                     cap["weekly"]["supply_min"], cap["weekly"]["utilization"], "total"])
    add_sheet(wb, "Capacity Model", ["Weekday", "Providers", "Demand (min)", "Supply (min)", "Utilization", "Flag"],
              cap_rows, note=f"Minute-weighted (ASK4). Demand from {cap['demand_source']}. Rebalance: {cap.get('rebalance') or '—'}",
              pct_cols=(5,), num_cols=(3, 4))
    index.append(("Capacity Model", len(cap_rows), "Minute-weighted demand vs supply by weekday (ASK4)."))

    grid = [[g["week_of"], g["Mon"], g["Tue"], g["Wed"], g["Thu"], g["Fri"], g["total"],
             "YES" if g["below_floor"] else "OK", g["util"]] for g in b["grid13"]]
    add_sheet(wb, "13-Week Grid", ["Week of", "Mon", "Tue", "Wed", "Thu", "Fri", "Wk total", "Below floor?", "Util %"],
              grid, note=f"Providers scheduled per weekday; floor = {b['floor']}/day.", pct_cols=(9,))
    index.append(("13-Week Grid", len(grid), "13-week staffing grid with below-floor flags."))

    canc = [[c["slot"], c["booked"], c["advance_pct"], c["noshow_pct"], c["cancel_pct"], c["signal"]]
            for c in b["cancellations"]]
    add_sheet(wb, "Cancellations", ["Slot", "Booked", "Advance %", "No-show %", "Cancel %", "Double-block signal"],
              canc, note="Advance (refilled) vs no-show (wasted) → double-block decision (ASK1).",
              pct_cols=(3, 4, 5))
    index.append(("Cancellations", len(canc), "Advance vs no-show by slot + double-block signal (ASK1)."))

    walk = [[w["weekday"], w["am"], w["pm"], w["total"], w["pm_share"], w["idle"], w["signal"]] for w in b["walkins"]]
    add_sheet(wb, "Walk-in Summary", ["Weekday", "AM", "PM", "Total/day", "PM share", "Idle risk (PM)", "Template signal"],
              walk, note="AM/PM split → full-day vs half-day walk-in template (ASK1).", pct_cols=(5,))
    index.append(("Walk-in Summary", len(walk), "Walk-in AM/PM split + template signal (ASK1)."))

    cyc = [[s["stage"], s["this_q"], s["last_q"], s["change"], s["pct"], s["flag"]] for s in b["cycle"]["stages"]]
    add_sheet(wb, "Cycle Time by Stage", ["Stage (owner)", "This Q (days)", "Last Q (days)", "Change", "% of total", "Flag"],
              cyc, note=f"Referral→seen by hand-off. Total {b['cycle']['total_this_q']}d vs {b['cycle']['total_last_q']}d; bottleneck: {b['cycle']['bottleneck_label']} (ASK3).",
              pct_cols=(5,))
    index.append(("Cycle Time by Stage", len(cyc), "Cycle time by hand-off with bottleneck attribution (ASK3)."))

    floors = [[f["service"], f["floor"]] for f in b["pto"]["floors"]]
    add_sheet(wb, "Coverage Floors", ["Service", "Min providers/day"], floors,
              note="Per-service minimum coverage floors checked on PTO approval (UC4/ASK2).")
    index.append(("Coverage Floors", len(floors), "Per-service minimum coverage floors (UC4/ASK2)."))

    reqs = [[r["week_of"], r["service"], r["on_floor"], r["if_approved"], r["result"], r["providers_out"], r["action"]]
            for r in b["pto"]["requests"]]
    add_sheet(wb, "Coverage Requests", ["Week", "Service", "On floor", "If approved", "Result", "Providers out", "Recommended action"],
              reqs, note=f"90-day forward PTO checks — {b['pto']['gap_count']} breach-day(s) (ASK2).")
    index.append(("Coverage Requests", len(reqs), "Forward PTO requests vs floors — breach results (ASK2)."))

    vt = [[v["type"], v["duration"], v["buffer"], v["total"], v["notes"]] for v in b["visit_types"]]
    add_sheet(wb, "Visit Types", ["Visit Type", "Duration (min)", "Buffer (min)", "Total (min)", "Notes"],
              vt, note="Reference durations that drive the capacity model.")
    index.append(("Visit Types", len(vt), "Visit-type durations feeding the capacity math."))

    _readme(wb, index)
    wb.save(OUT)
    print(f"wrote {OUT}  ({OUT.stat().st_size} bytes, {len(wb.sheetnames)} sheets)")
    print("sheets:", ", ".join(wb.sheetnames))


if __name__ == "__main__":
    build()
